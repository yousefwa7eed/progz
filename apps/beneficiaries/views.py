from django.views.generic import ListView, DetailView, CreateView, UpdateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django import forms
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment
from fpdf import FPDF
import arabic_reshaper
from bidi.algorithm import get_display
from .models import Beneficiary
from apps.cases.models import Case
from apps.donations.models import Donation
from apps.documents.models import Document


class BeneficiaryForm(forms.ModelForm):
    class Meta:
        model = Beneficiary
        fields = ['full_name', 'gender', 'national_id', 'birth_date', 'phone', 'phone2',
                  'address', 'city', 'district', 'marital_status', 'family_members',
                  'has_orphans', 'orphans_count', 'health_status', 'has_chronic_disease',
                  'chronic_diseases', 'has_disabilities', 'disabilities_details',
                  'employment_status', 'monthly_income', 'housing_type', 'is_urgent', 'notes']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['family_members'].required = False
        self.fields['orphans_count'].required = False
        self.fields['monthly_income'].required = False
        self.fields['notes'].required = False


class BeneficiaryListView(LoginRequiredMixin, ListView):
    model = Beneficiary
    template_name = 'beneficiaries/list.html'
    context_object_name = 'beneficiaries'
    paginate_by = 25

    def get_queryset(self):
        qs = Beneficiary.objects.select_related('branch').filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(Q(full_name__icontains=search) | Q(phone__icontains=search) |
                           Q(code__icontains=search) | Q(national_id__icontains=search))
        city = self.request.GET.get('city')
        if city:
            qs = qs.filter(city=city)
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        urgent = self.request.GET.get('urgent')
        if urgent:
            qs = qs.filter(is_urgent=True)
        return qs.order_by('-priority_score', '-created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['cities'] = Beneficiary.objects.values_list('city', flat=True).distinct()
        ctx['search'] = self.request.GET.get('search', '')
        ctx['city_filter'] = self.request.GET.get('city', '')
        ctx['status_filter'] = self.request.GET.get('status', '')
        ctx['urgent_filter'] = self.request.GET.get('urgent', '')
        return ctx


class BeneficiaryCreateView(LoginRequiredMixin, CreateView):
    model = Beneficiary
    form_class = BeneficiaryForm
    template_name = 'beneficiaries/form.html'
    success_url = '/beneficiaries/'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'تم إضافة المستفيد بنجاح')
        return super().form_valid(form)


class BeneficiaryDetailView(LoginRequiredMixin, DetailView):
    model = Beneficiary
    template_name = 'beneficiaries/detail.html'
    context_object_name = 'beneficiary'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['cases'] = Case.objects.filter(beneficiary=self.object).order_by('-created_at')
        ctx['donations'] = Donation.objects.filter(
            items__contains=[{'beneficiary_id': str(self.object.id)}]
        ).order_by('-created_at')[:10]
        ctx['documents'] = Document.objects.filter(
            related_model='Beneficiary', related_id=self.object.id
        ).order_by('-created_at')
        return ctx


class BeneficiaryUpdateView(LoginRequiredMixin, UpdateView):
    model = Beneficiary
    form_class = BeneficiaryForm
    template_name = 'beneficiaries/form.html'
    success_url = '/beneficiaries/'

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث بيانات المستفيد')
        return super().form_valid(form)


@login_required
def beneficiary_search(request):
    q = request.GET.get('q', '')
    beneficiaries = Beneficiary.objects.filter(
        Q(full_name__icontains=q) | Q(phone__icontains=q) | Q(code__icontains=q),
        is_active=True
    )[:20]
    data = [{'id': str(b.id), 'text': f'{b.code} - {b.full_name}'} for b in beneficiaries]
    return JsonResponse({'results': data})


@login_required
def beneficiary_quick_add(request):
    import json
    if request.method == 'POST':
        name = request.POST.get('full_name', '').strip()
        gender = request.POST.get('gender', '')
        phone = request.POST.get('phone', '').strip()
        if not name or not gender or not phone:
            return JsonResponse({'error': 'الاسم والجنس ورقم الموبايل مطلوبة'}, status=400)
        b = Beneficiary(full_name=name, gender=gender, phone=phone, created_by=request.user)
        b.save()
        return JsonResponse({'id': str(b.id), 'code': b.code, 'name': b.full_name})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def beneficiary_delete(request, pk):
    b = get_object_or_404(Beneficiary, pk=pk)
    if request.method == 'POST':
        b.is_active = False
        b.deleted_at = timezone.now()
        b.save()
        messages.warning(request, f'تم حذف المستفيد {b.full_name}')
        return redirect('beneficiary_list')
    return render(request, 'beneficiaries/confirm_delete.html', {'object': b, 'title': 'حذف مستفيد'})


@login_required
def beneficiary_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المستفيدين'
    headers = ['الكود', 'الاسم', 'الجنس', 'رقم الموبايل', 'المحافظة', 'الحالة الاجتماعية',
               'عدد الأسرة', 'الدخل', 'الحالة', 'درجة الأولوية', 'تاريخ التسجيل']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row, b in enumerate(Beneficiary.objects.filter(is_active=True), 2):
        vals = [b.code, b.full_name, b.get_gender_display(), b.phone, b.city,
                b.get_marital_status_display() if b.marital_status else '',
                b.family_members, b.monthly_income, b.get_status_display(),
                b.priority_score, b.created_at.strftime('%Y-%m-%d')]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=beneficiaries.xlsx'
    wb.save(response)
    return response


@login_required
def beneficiary_export_detail_excel(request, pk):
    b = get_object_or_404(Beneficiary, pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ملف المستفيد'
    ws.cell(row=1, column=1, value='ملف المستفيد: ' + b.full_name).font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    headers = ['المعلومة', 'القيمة']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    data = [
        ('الكود', b.code), ('الاسم', b.full_name), ('الجنس', b.get_gender_display()),
        ('الرقم القومي', b.national_id or '--'), ('رقم الموبايل', b.phone),
        ('المحافظة', b.city or '--'), ('المنطقة', b.district or '--'),
        ('العنوان', b.address or '--'), ('الحالة الاجتماعية', b.get_marital_status_display() or '--'),
        ('عدد الأسرة', str(b.family_members) if b.family_members else '--'),
        ('الدخل الشهري', str(b.monthly_income) if b.monthly_income else '--'),
        ('الحالة', b.get_status_display()),
    ]
    for row, (k, v) in enumerate(data, 4):
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=beneficiary_{b.code}.xlsx'
    wb.save(response)
    return response


@login_required
def beneficiary_export_detail_pdf(request, pk):
    b = get_object_or_404(Beneficiary, pk=pk)
    import os

    def ar(text):
        return get_display(arabic_reshaper.reshape(str(text)))

    font_path = bold_path = 'C:/Windows/Fonts/arial.ttf'
    for fp in ['C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/arabtype.ttf',
               'C:/Windows/Fonts/trado.ttf']:
        if os.path.exists(fp):
            font_path = fp
            bold_path = fp
            if 'arial' in fp.lower():
                bold_path = 'C:/Windows/Fonts/arialbd.ttf'
            break

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=28)
    pdf.add_font('ArFont', '', font_path)
    pdf.add_font('ArFont', 'B', bold_path)
    pdf.add_page()

    # ── Full-page border ──
    pdf.set_draw_color(0, 51, 102)
    pdf.set_line_width(1)
    pdf.rect(5, 5, pdf.w - 10, pdf.h - 10)
    pdf.set_draw_color(210, 218, 235)
    pdf.set_line_width(0.3)
    pdf.rect(8, 8, pdf.w - 16, pdf.h - 16)

    # ── Header: Logo + name ──
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
    if os.path.exists(logo_path):
        pdf.image(logo_path, x=16, y=16, w=20)
    pdf.set_font('ArFont', 'B', 20)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(0, 11, ar('مؤسسة الماجد'), new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.set_font('ArFont', '', 10)
    pdf.set_text_color(110, 110, 110)
    pdf.cell(0, 6, ar('الجهاز الإداري'), new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.ln(3)

    # ── Separator line ──
    pdf.set_draw_color(0, 51, 102)
    pdf.set_line_width(0.4)
    pdf.line(14, pdf.get_y(), pdf.w - 14, pdf.get_y())
    pdf.ln(6)

    # ── Title + date ──
    pdf.set_font('ArFont', 'B', 15)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(0, 10, ar('ملف المستفيد'), new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.set_font('ArFont', '', 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, ar(f'تاريخ الطباعة: {timezone.now().strftime("%Y/%m/%d %H:%M")}'), new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.ln(5)

    def section(title):
        pdf.ln(3)
        pdf.set_fill_color(0, 51, 102)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('ArFont', 'B', 12)
        pdf.cell(0, 9, ar(f'  {title}'), border=0, align='R', fill=True)
        pdf.ln(5)
        pdf.set_text_color(30, 30, 30)

    def data_table(rows):
        rh = 9
        pdf.set_draw_color(200, 210, 230)
        for i, (lbl, val) in enumerate(rows):
            fill = i % 2 == 0
            pdf.set_fill_color(245, 248, 255) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.set_font('ArFont', '', 10)
            pdf.set_text_color(70, 70, 70)
            pdf.cell(52, rh, ar(f' {lbl}'), border='LR', align='R', fill=True)
            pdf.set_font('ArFont', 'B', 11)
            pdf.set_text_color(10, 10, 10)
            pdf.cell(0, rh, ar(f'  {val}'), border='LR', align='R', fill=True, new_x='LMARGIN', new_y='NEXT')
        pdf.cell(0, 2, '', border='T', new_x='LMARGIN', new_y='NEXT')

    # ── Basic info section ──
    section('البيانات الأساسية')
    data_table([
        ('الاسم', b.full_name),
        ('الكود', b.code),
        ('الجنس', b.get_gender_display()),
        ('الرقم القومي', b.national_id or '--'),
        ('رقم الموبايل', b.phone),
        ('المحافظة', b.city or '--'),
        ('المنطقة', b.district or '--'),
        ('العنوان', b.address or '--'),
    ])

    # ── Social & financial info ──
    section('البيانات الاجتماعية والمالية')
    data_table([
        ('الحالة الاجتماعية', b.get_marital_status_display() or '--'),
        ('عدد أفراد الأسرة', str(b.family_members) if b.family_members else '--'),
        ('الدخل الشهري', (f'{b.monthly_income:,.2f} جنيه') if b.monthly_income else '--'),
        ('الحالة', b.get_status_display()),
        ('درجة الأولوية', str(b.priority_score)),
    ])

    # ── Cases table ──
    cases = Case.objects.filter(beneficiary=b).order_by('-created_at')
    if cases.exists():
        section('الحالات السابقة')
        pdf.set_draw_color(200, 210, 230)
        # Header row
        pdf.set_fill_color(0, 51, 102)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('ArFont', 'B', 11)
        col_w = [50, 45, 45, 50]
        pdf.cell(col_w[0], 9, ar('  نوع الحالة'), border=1, align='R', fill=True)
        pdf.cell(col_w[1], 9, ar('الحالة'), border=1, align='C', fill=True)
        pdf.cell(col_w[2], 9, ar('المبلغ'), border=1, align='C', fill=True)
        pdf.cell(col_w[3], 9, ar('التاريخ'), border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(30, 30, 30)
        for i, c in enumerate(cases):
            fill = i % 2 == 0
            pdf.set_fill_color(245, 248, 255) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.set_font('ArFont', '', 10)
            pdf.cell(col_w[0], 9, ar(f'  {c.get_case_type_display()}'), border=1, align='R', fill=fill)
            pdf.cell(col_w[1], 9, ar(c.get_status_display()), border=1, align='C', fill=fill)
            pdf.cell(col_w[2], 9, ar(f'{c.requested_amount:,.0f}') if c.requested_amount else '--', border=1, align='C', fill=fill)
            pdf.cell(col_w[3], 9, ar(c.created_at.strftime('%Y-%m-%d')), border=1, align='C', fill=fill)
            pdf.ln()

    # ── Stamp circle at bottom ──
    pdf.ln(12)
    if pdf.get_y() > pdf.h - 50:
        pdf.add_page()
    r = 15
    cx = pdf.w / 2
    cy = pdf.get_y() + r + 3
    pdf.set_draw_color(0, 51, 102)
    pdf.set_line_width(0.6)
    pdf.circle(cx, cy, r)
    pdf.set_y(cy + r + 5)
    pdf.set_font('ArFont', '', 9)
    pdf.set_text_color(110, 110, 110)
    pdf.cell(0, 6, ar('ختم الإدارة'), align='C')

    return HttpResponse(bytes(pdf.output()), content_type='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename=beneficiary_{b.code}.pdf'})
