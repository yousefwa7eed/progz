from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment
from .models import Donor, DonorCategory
from apps.donations.models import Donation
from apps.sponsorships.models import Sponsorship


class DonorListView(LoginRequiredMixin, ListView):
    model = Donor
    template_name = 'donors/list.html'
    context_object_name = 'donors'
    paginate_by = 25

    def get_queryset(self):
        qs = Donor.objects.select_related('donor_category').filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(Q(full_name__icontains=search) | Q(phone__icontains=search) |
                           Q(code__icontains=search) | Q(email__icontains=search))
        donor_type = self.request.GET.get('donor_type')
        if donor_type:
            qs = qs.filter(donor_type=donor_type)
        committed = self.request.GET.get('committed')
        if committed:
            qs = qs.filter(is_committed=True)
        return qs.order_by('-total_donations')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search'] = self.request.GET.get('search', '')
        ctx['type_filter'] = self.request.GET.get('donor_type', '')
        ctx['committed_filter'] = self.request.GET.get('committed', '')
        return ctx


class DonorCreateView(LoginRequiredMixin, CreateView):
    model = Donor
    template_name = 'donors/form.html'
    fields = ['full_name', 'donor_type', 'phone', 'email', 'address', 'city',
              'national_id', 'commercial_reg', 'contact_person',
              'preferred_contact', 'preferred_donation', 'is_anonymous',
              'donor_category', 'notes']
    success_url = '/donors/'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'تم إضافة المتبرع بنجاح')
        return super().form_valid(form)


class DonorDetailView(LoginRequiredMixin, DetailView):
    model = Donor
    template_name = 'donors/detail.html'
    context_object_name = 'donor'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['donations'] = Donation.objects.filter(donor=self.object).order_by('-created_at')[:20]
        ctx['sponsorships'] = Sponsorship.objects.filter(sponsor=self.object).order_by('-created_at')
        return ctx


class DonorUpdateView(LoginRequiredMixin, UpdateView):
    model = Donor
    template_name = 'donors/form.html'
    fields = ['full_name', 'donor_type', 'phone', 'email', 'address', 'city',
              'national_id', 'commercial_reg', 'contact_person',
              'preferred_contact', 'preferred_donation', 'is_anonymous',
              'donor_category', 'notes']
    success_url = '/donors/'

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث بيانات المتبرع')
        return super().form_valid(form)


@login_required
def donor_delete(request, pk):
    d = get_object_or_404(Donor, pk=pk)
    if request.method == 'POST':
        d.is_active = False
        d.save()
        messages.warning(request, f'تم حذف المتبرع {d.full_name}')
        return redirect('donor_list')
    return render(request, 'beneficiaries/confirm_delete.html', {'object': d, 'title': 'حذف متبرع'})


@login_required
def donor_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المتبرعين'
    headers = ['الاسم', 'النوع', 'رقم الموبايل', 'البريد', 'المحافظة', 'إجمالي التبرعات', 'آخر تبرع']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row, d in enumerate(Donor.objects.filter(is_active=True), 2):
        vals = [d.full_name, d.get_donor_type_display(), d.phone, d.email or '',
                d.city or '', float(d.total_donations),
                d.last_donation_date.strftime('%Y-%m-%d') if d.last_donation_date else '']
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=donors.xlsx'
    wb.save(response)
    return response
