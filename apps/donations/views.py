from django.views.generic import ListView, DetailView, CreateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment
from .models import Donation, DonationItem
from apps.finance.models import FinancialEntry, Account


class DonationListView(LoginRequiredMixin, ListView):
    model = Donation
    template_name = 'donations/list.html'
    context_object_name = 'donations'
    paginate_by = 25

    def get_queryset(self):
        qs = Donation.objects.select_related('donor', 'project', 'received_by').all()
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(Q(code__icontains=search) | Q(receipt_number__icontains=search) |
                           Q(donor__full_name__icontains=search) | Q(donor_name__icontains=search))
        donation_type = self.request.GET.get('donation_type')
        if donation_type:
            qs = qs.filter(donation_type=donation_type)
        transaction_type = self.request.GET.get('transaction_type')
        if transaction_type:
            qs = qs.filter(transaction_type=transaction_type)
        date_from = self.request.GET.get('date_from')
        if date_from:
            qs = qs.filter(receipt_date__gte=date_from)
        date_to = self.request.GET.get('date_to')
        if date_to:
            qs = qs.filter(receipt_date__lte=date_to)
        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search'] = self.request.GET.get('search', '')
        ctx['type_filter'] = self.request.GET.get('donation_type', '')
        ctx['txn_filter'] = self.request.GET.get('transaction_type', '')
        ctx['date_from'] = self.request.GET.get('date_from', '')
        ctx['date_to'] = self.request.GET.get('date_to', '')
        ctx['total_amount'] = self.get_queryset().aggregate(total=Sum('amount'))['total'] or 0
        return ctx


class DonationCreateView(LoginRequiredMixin, CreateView):
    model = Donation
    template_name = 'donations/form.html'
    fields = ['donor', 'donor_name', 'donation_type', 'payment_method', 'amount',
              'transaction_type', 'is_zakat', 'zakat_year', 'project', 'notes']
    success_url = '/donations/'

    def form_valid(self, form):
        form.instance.received_by = self.request.user
        form.instance.created_by = self.request.user
        form.instance.status = 'confirmed'
        response = super().form_valid(form)

        # Create financial entry automatically
        acct_map = {'general': '3100', 'zakat': '3200', 'sadaqah': '3300', 'sponsorship': '3300', 'project': '3300', 'endowment': '3300'}
        acct_code = acct_map.get(form.instance.transaction_type, '3300')
        account = Account.objects.filter(code=acct_code).first()
        if account:
            FinancialEntry.objects.create(
                entry_type='income',
                entry_date=form.instance.receipt_date,
                amount=form.instance.amount,
                description=f'تبرع {form.instance.get_transaction_type_display()} - {form.instance.code}',
                account=account,
                donor=form.instance.donor,
                donation=form.instance,
                payment_method=form.instance.payment_method,
                receipt_number=form.instance.receipt_number,
                is_approved=True,
                recorded_by=self.request.user,
                transaction_type=form.instance.transaction_type,
            )

        messages.success(self.request, f'تم تسجيل التبرع بنجاح. رقم الإيصال: {form.instance.receipt_number}')
        return response


class DonationDetailView(LoginRequiredMixin, DetailView):
    model = Donation
    template_name = 'donations/detail.html'
    context_object_name = 'donation'


@login_required
def donation_receipt(request, pk):
    donation = get_object_or_404(Donation, pk=pk)
    return render(request, 'donations/receipt.html', {'donation': donation})


@login_required
def donation_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'التبرعات'
    headers = ['رقم الإيصال', 'المتبرع', 'النوع', 'المبلغ', 'طريقة الدفع', 'التاريخ']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row, d in enumerate(Donation.objects.select_related('donor').all(), 2):
        vals = [d.receipt_number, d.donor.full_name if d.donor and not d.is_anonymous else 'متبرع',
                d.get_donation_type_display(), float(d.amount), d.get_payment_method_display(),
                d.receipt_date.strftime('%Y-%m-%d') if d.receipt_date else '']
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=donations.xlsx'
    wb.save(response)
    return response
