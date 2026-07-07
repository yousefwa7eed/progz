import uuid
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models, transaction
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment
from fpdf import FPDF
import arabic_reshaper
from bidi.algorithm import get_display
from .models import Occasion, OccasionMember, OccasionTask
from .forms import OccasionForm, BulkMemberAddForm, TaskForm
from apps.cases.models import Case
from apps.beneficiaries.models import Beneficiary


@login_required
def occasion_list(request):
    occasions = Occasion.objects.select_related('created_by').all()
    return render(request, 'occasions/list.html', {
        'occasions': occasions,
        'active_count': occasions.filter(status='active').count(),
        'total_count': occasions.count(),
    })


@login_required
def occasion_create(request):
    if request.method == 'POST':
        form = OccasionForm(request.POST)
        if form.is_valid():
            occasion = form.save(commit=False)
            occasion.created_by = request.user
            occasion.save()
            messages.success(request, 'تم إنشاء المناسبة بنجاح')
            return redirect('occasion_detail', pk=occasion.id)
    else:
        form = OccasionForm()
    return render(request, 'occasions/form.html', {'form': form, 'title': 'إضافة مناسبة جديدة'})


@login_required
def occasion_edit(request, pk):
    occasion = get_object_or_404(Occasion, pk=pk)
    if request.method == 'POST':
        form = OccasionForm(request.POST, instance=occasion)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث المناسبة بنجاح')
            return redirect('occasion_detail', pk=occasion.id)
    else:
        form = OccasionForm(instance=occasion)
    return render(request, 'occasions/form.html', {'form': form, 'occasion': occasion, 'title': f'تعديل: {occasion.name}'})


@login_required
def occasion_detail(request, pk):
    occasion = get_object_or_404(Occasion.objects.prefetch_related('members__tasks', 'members__case__beneficiary', 'members__beneficiary'), pk=pk)
    if request.method == 'POST':
        member_ids = request.POST.getlist('bulk_complete')
        if member_ids:
            occasion.members.filter(pk__in=member_ids).update(completed=True)
            messages.success(request, f'تم إنجاز {len(member_ids)} أعضاء بنجاح')
        return redirect('occasion_detail', pk=occasion.id)
    members = occasion.members.all()
    completed_members = members.filter(completed=True).count()
    total_members = members.count()
    progress = int((completed_members / total_members * 100)) if total_members > 0 else 0
    return render(request, 'occasions/detail.html', {
        'occasion': occasion,
        'members': members,
        'completed_members': completed_members,
        'total_members': total_members,
        'progress': progress,
    })


@login_required
def occasion_delete(request, pk):
    occasion = get_object_or_404(Occasion, pk=pk)
    if request.method == 'POST':
        occasion.delete()
        messages.success(request, 'تم حذف المناسبة بنجاح')
        return redirect('occasion_list')
    return render(request, 'occasions/confirm_delete.html', {'occasion': occasion})


@login_required
def occasion_add_members(request, pk):
    occasion = get_object_or_404(Occasion, pk=pk)
    if request.method == 'POST':
        member_type = request.POST.get('member_type')
        notes = request.POST.get('notes', '')
        added_count = 0

        if member_type == 'case':
            case_ids = request.POST.getlist('case_ids')
            for cid in case_ids:
                try:
                    case = Case.objects.get(pk=uuid.UUID(cid))
                    OccasionMember.objects.get_or_create(
                        occasion=occasion,
                        member_type='case',
                        case=case,
                        defaults={'beneficiary': case.beneficiary, 'notes': notes, 'added_by': request.user}
                    )
                    added_count += 1
                except (ValueError, Case.DoesNotExist):
                    pass
        else:
            ben_ids = request.POST.getlist('beneficiary_ids')
            for bid in ben_ids:
                try:
                    beneficiary = Beneficiary.objects.get(pk=uuid.UUID(bid))
                    OccasionMember.objects.get_or_create(
                        occasion=occasion,
                        member_type='beneficiary',
                        beneficiary=beneficiary,
                        defaults={'notes': notes, 'added_by': request.user}
                    )
                    added_count += 1
                except (ValueError, Beneficiary.DoesNotExist):
                    pass

        messages.success(request, f'تم إضافة {added_count} أعضاء بنجاح')
        return redirect('occasion_detail', pk=occasion.id)

    cases = Case.objects.select_related('beneficiary').exclude(status__in=['closed', 'rejected']).order_by('-created_at')
    beneficiaries = Beneficiary.objects.filter(is_active=True, status='active').order_by('-created_at')
    return render(request, 'occasions/add_members.html', {
        'occasion': occasion,
        'cases': cases,
        'beneficiaries': beneficiaries,
    })


@login_required
def occasion_remove_member(request, pk, member_pk):
    member = get_object_or_404(OccasionMember, pk=member_pk, occasion_id=pk)
    if request.method == 'POST':
        member.delete()
        messages.success(request, 'تم حذف العضو بنجاح')
    return redirect('occasion_detail', pk=pk)


@login_required
def member_toggle_complete(request, pk, member_pk):
    member = get_object_or_404(OccasionMember, pk=member_pk, occasion_id=pk)
    member.completed = not member.completed
    member.save()
    return redirect('occasion_detail', pk=pk)


@login_required
def task_add(request, pk, member_pk):
    member = get_object_or_404(OccasionMember, pk=member_pk, occasion_id=pk)
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.member = member
            task.save()
            messages.success(request, 'تم إضافة المهمة بنجاح')
    return redirect('occasion_detail', pk=pk)


@login_required
def task_toggle(request, pk, member_pk, task_pk):
    task = get_object_or_404(OccasionTask, pk=task_pk, member_id=member_pk, member__occasion_id=pk)
    if task.status == 'completed':
        task.status = 'pending'
        task.completed_at = None
        task.completed_by = None
    else:
        task.status = 'completed'
        task.completed_at = timezone.now()
        task.completed_by = request.user
    task.save()
    return redirect('occasion_detail', pk=pk)


@login_required
def task_delete(request, pk, member_pk, task_pk):
    task = get_object_or_404(OccasionTask, pk=task_pk, member_id=member_pk, member__occasion_id=pk)
    task.delete()
    messages.success(request, 'تم حذف المهمة')
    return redirect('occasion_detail', pk=pk)


@login_required
def occasion_export_pdf(request, pk):
    occasion = get_object_or_404(Occasion.objects.prefetch_related('members__tasks', 'members__case__beneficiary', 'members__beneficiary'), pk=pk)
    import os

    def ar(text):
        return get_display(arabic_reshaper.reshape(str(text)))

    font_path = 'C:/Windows/Fonts/arial.ttf'
    bold_path = 'C:/Windows/Fonts/arialbd.ttf'
    for fp in ['C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/arabtype.ttf',
               'C:/Windows/Fonts/trado.ttf']:
        if os.path.exists(fp):
            font_path = fp
            bold_path = fp
            if 'arial' in fp.lower():
                bold_path = 'C:/Windows/Fonts/arialbd.ttf'
            break

    class PDFWithHeader(FPDF):
        def header(self):
            self.set_fill_color(85, 107, 47)
            self.rect(5, 5, 200, 287, 'D')
            self.rect(8, 8, 194, 281, 'D')
            self.set_y(15)
            logo = os.path.join('static', 'images', 'logo.png')
            if os.path.exists(logo):
                self.image(logo, x=85, w=40, h=0)
            self.set_y(28)
            self.set_font('ArFont', 'B', 16)
            self.cell(0, 10, ar('مؤسسة الماجد'), align='C')
            self.ln(8)
            self.set_font('ArFont', '', 13)
            self.cell(0, 10, ar(occasion.name), align='C')
            self.ln(6)
            self.set_font('ArFont', '', 10)
            self.cell(0, 8, ar(f"{occasion.display_support_type} | {timezone.now().strftime('%Y/%m/%d')}"), align='C')
            self.ln(10)

        def footer(self):
            self.set_y(-15)
            self.set_font('ArFont', '', 8)
            self.cell(0, 10, ar(f'صفحة {self.page_no()}'), align='C')

    pdf = PDFWithHeader('P', 'mm', 'A4')
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_font('ArFont', '', font_path)
    pdf.add_font('ArFont', 'B', bold_path)
    pdf.add_page()

    col_widths = [10, 60, 20, 20, 80]
    headers = ['م', 'الاسم', 'النوع', 'الحالة', 'المهام']

    pdf.set_font('ArFont', 'B', 10)
    pdf.set_fill_color(85, 107, 47)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, ar(h), border=1, align='C', fill=True)
    pdf.ln()

    pdf.set_font('ArFont', '', 9)
    pdf.set_text_color(0, 0, 0)
    for idx, m in enumerate(occasion.members.all(), 1):
        fill = idx % 2 == 0
        if fill:
            pdf.set_fill_color(249, 249, 249)
        pdf.cell(col_widths[0], 7, str(idx), border=1, align='C', fill=fill)
        pdf.cell(col_widths[1], 7, ar(m.display_name), border=1, align='C', fill=fill)
        pdf.cell(col_widths[2], 7, ar(m.get_member_type_display()), border=1, align='C', fill=fill)
        pdf.cell(col_widths[3], 7, ar('تم' if m.completed else 'قيد الانتظار'), border=1, align='C', fill=fill)
        tasks_text = ' - '.join([t.task_name for t in m.tasks.all()]) or '-'
        pdf.cell(col_widths[4], 7, ar(tasks_text[:60]), border=1, align='C', fill=fill)
        pdf.ln()

    return HttpResponse(bytes(pdf.output()), content_type='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{occasion.name}.pdf"'})


@login_required
def occasion_export_excel(request, pk):
    occasion = get_object_or_404(Occasion.objects.prefetch_related('members__tasks', 'members__case__beneficiary', 'members__beneficiary'), pk=pk)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'الأعضاء'
    headers = ['الاسم', 'النوع', 'الحالة', 'عدد المهام', 'المهام المنجزة', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for i, m in enumerate(occasion.members.all(), 2):
        ws.cell(row=i, column=1, value=m.display_name)
        ws.cell(row=i, column=2, value=dict(OccasionMember.MEMBER_TYPES).get(m.member_type, ''))
        ws.cell(row=i, column=3, value='تم' if m.completed else 'قيد الانتظار')
        ws.cell(row=i, column=4, value=m.tasks.count())
        ws.cell(row=i, column=5, value=m.completed_tasks_count)
        ws.cell(row=i, column=6, value=m.notes or '')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{occasion.name}.xlsx"'
    wb.save(response)
    return response


@login_required
def api_search_cases(request):
    q = request.GET.get('q', '')
    cases = Case.objects.select_related('beneficiary').filter(status__in=['new', 'under_study', 'approved'])
    if q:
        cases = cases.filter(
            models.Q(code__icontains=q) |
            models.Q(beneficiary__full_name__icontains=q) |
            models.Q(beneficiary__phone__icontains=q)
        )
    results = [{
        'id': str(c.id),
        'text': f'{c.beneficiary.full_name} ({c.code}) - {c.get_case_type_display()}',
    } for c in cases[:50]]
    return JsonResponse({'results': results})


@login_required
def api_search_beneficiaries(request):
    q = request.GET.get('q', '')
    beneficiaries = Beneficiary.objects.filter(is_active=True)
    if q:
        beneficiaries = beneficiaries.filter(
            models.Q(full_name__icontains=q) |
            models.Q(phone__icontains=q) |
            models.Q(national_id__icontains=q)
        )
    results = [{
        'id': str(b.id),
        'text': f'{b.full_name} - {b.phone or ""}',
    } for b in beneficiaries[:50]]
    return JsonResponse({'results': results})
