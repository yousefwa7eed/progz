from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
import json
import openpyxl
from openpyxl.styles import Font, Alignment
from .models import Case, CaseActivity, CaseImage, CaseFeature
from .forms import CaseForm


class CaseListView(LoginRequiredMixin, ListView):
    model = Case
    template_name = 'cases/list.html'
    context_object_name = 'cases'
    paginate_by = 25

    def get_queryset(self):
        qs = Case.objects.select_related('beneficiary', 'assigned_to').all()
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(Q(code__icontains=search) | Q(beneficiary__full_name__icontains=search) |
                           Q(description__icontains=search))
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        priority = self.request.GET.get('priority')
        if priority:
            qs = qs.filter(priority=priority)
        case_type = self.request.GET.get('case_type')
        if case_type:
            qs = qs.filter(case_type=case_type)
        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search'] = self.request.GET.get('search', '')
        ctx['status_filter'] = self.request.GET.get('status', '')
        ctx['priority_filter'] = self.request.GET.get('priority', '')
        ctx['type_filter'] = self.request.GET.get('case_type', '')
        return ctx


class CaseCreateView(LoginRequiredMixin, CreateView):
    model = Case
    form_class = CaseForm
    template_name = 'cases/form.html'
    success_url = '/cases/'

    def get_initial(self):
        initial = super().get_initial()
        beneficiary_id = self.request.GET.get('beneficiary')
        if beneficiary_id:
            from apps.beneficiaries.models import Beneficiary
            try:
                initial['beneficiary'] = Beneficiary.objects.get(id=beneficiary_id, is_active=True)
            except Beneficiary.DoesNotExist:
                messages.warning(self.request, 'المستفيد المطلوب غير موجود')
        return initial

    def get_success_url(self):
        case_type = self.object.case_type
        if case_type in dict(Case.CASE_TYPES):
            return f'/cases/type/{case_type}/'
        return '/cases/'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        features_json = self.request.POST.get('features_json', '[]')
        try:
            features = json.loads(features_json)
            for feat in features:
                name = feat.get('name', '').strip()
                value = feat.get('value', '').strip()
                if name and value:
                    CaseFeature.objects.create(case=self.object, name=name, value=value)
        except json.JSONDecodeError:
            messages.warning(self.request, 'لم نتمكن من قراءة الخصائص المدخلة، يرجى التحقق من التنسيق')
        CaseActivity.objects.create(
            case=self.object,
            activity_type='create',
            description='تم إنشاء الحالة',
            performed_by=self.request.user,
            new_status='new',
        )
        images = self.request.FILES.getlist('images')
        for i, img in enumerate(images):
            labels = self.request.POST.getlist('img_labels')
            label = labels[i] if i < len(labels) and labels[i].strip() else 'صورة'
            CaseImage.objects.create(case=self.object, image=img, label=label)
        messages.success(self.request, 'تم إنشاء الحالة بنجاح')
        return response


class CaseDetailView(LoginRequiredMixin, DetailView):
    model = Case
    template_name = 'cases/detail.html'
    context_object_name = 'case'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['activities'] = CaseActivity.objects.filter(case=self.object).order_by('-created_at')
        ctx['images'] = CaseImage.objects.filter(case=self.object).order_by('-uploaded_at')
        ctx['features'] = CaseFeature.objects.filter(case=self.object).order_by('name')
        return ctx


@login_required
@require_POST
def case_approve(request, pk):
    case = get_object_or_404(Case, pk=pk)
    if case.status != 'under_study':
        messages.error(request, 'لا يمكن اعتماد الحالة في حالتها الحالية')
        return redirect('case_detail', pk=pk)
    amount = request.POST.get('approved_amount')
    case.status = 'approved'
    case.approved_amount = amount or case.requested_amount
    case.approved_by = request.user
    case.approved_at = timezone.now()
    case.save()
    CaseActivity.objects.create(
        case=case, activity_type='approve', performed_by=request.user,
        description=f'تم اعتماد الحالة بمبلغ {case.approved_amount}',
        old_status='under_study', new_status='approved',
    )
    messages.success(request, 'تم اعتماد الحالة')
    return redirect('case_detail', pk=pk)


@login_required
@require_POST
def case_reject(request, pk):
    case = get_object_or_404(Case, pk=pk)
    if case.status != 'under_study':
        messages.error(request, 'لا يمكن رفض الحالة في حالتها الحالية')
        return redirect('case_detail', pk=pk)
    reason = request.POST.get('reject_reason', '')
    case.status = 'rejected'
    case.close_reason = reason
    case.save()
    CaseActivity.objects.create(
        case=case, activity_type='reject', performed_by=request.user,
        description=f'تم رفض الحالة: {reason}',
        old_status='under_study', new_status='rejected',
    )
    messages.warning(request, 'تم رفض الحالة')
    return redirect('case_detail', pk=pk)


@login_required
@require_POST
def case_disburse(request, pk):
    case = get_object_or_404(Case, pk=pk)
    if case.status != 'approved':
        messages.error(request, 'لا يمكن صرف الحالة في حالتها الحالية')
        return redirect('case_detail', pk=pk)
    case.status = 'disbursed'
    case.save()
    from apps.finance.models import FinancialEntry, Account
    acct = Account.objects.filter(code='4100').first()
    if acct and case.approved_amount:
        FinancialEntry.objects.create(
            entry_type='expense', entry_date=timezone.now().date(),
            amount=case.approved_amount,
            description=f'صرف مساعدة - {case.code} - {case.beneficiary.full_name}',
            account=acct, case=case,
            is_approved=True, recorded_by=request.user,
            transaction_type='general',
        )
    CaseActivity.objects.create(
        case=case, activity_type='disburse', performed_by=request.user,
        description=f'تم صرف المساعدة',
        old_status='approved', new_status='disbursed',
    )
    messages.success(request, 'تم صرف المساعدة')
    return redirect('case_detail', pk=pk)


@login_required
def case_delete(request, pk):
    case = get_object_or_404(Case, pk=pk)
    if request.method == 'POST':
        case.is_active = False
        case.save()
        messages.warning(request, f'تم حذف الحالة {case.code}')
        return redirect('case_list')
    return render(request, 'beneficiaries/confirm_delete.html', {'object': case, 'title': 'حذف حالة'})


@login_required
def case_add_image(request, pk):
    case = get_object_or_404(Case, pk=pk)
    if request.method == 'POST':
        image_file = request.FILES.get('image')
        label = request.POST.get('label', '')
        if image_file:
            CaseImage.objects.create(case=case, image=image_file, label=label or 'صورة')
            messages.success(request, 'تم إضافة الصورة')
        else:
            messages.error(request, 'يرجى اختيار صورة')
    return redirect('case_detail', pk=pk)


@login_required
def case_delete_image(request, pk, image_id):
    img = get_object_or_404(CaseImage, id=image_id, case_id=pk)
    img.delete()
    messages.success(request, 'تم حذف الصورة')
    return redirect('case_detail', pk=pk)


@login_required
@require_POST
def case_add_feature(request, pk):
    case = get_object_or_404(Case, pk=pk)
    name = request.POST.get('name', '').strip()
    value = request.POST.get('value', '').strip()
    if name and value:
        CaseFeature.objects.create(case=case, name=name, value=value)
        messages.success(request, f'تم إضافة الخاصية {name}')
    else:
        messages.error(request, 'يرجى إدخال اسم الخاصية والقيمة')
    return redirect('case_detail', pk=pk)


@login_required
def case_delete_feature(request, pk, feature_id):
    feat = get_object_or_404(CaseFeature, id=feature_id, case_id=pk)
    feat.delete()
    messages.success(request, 'تم حذف الخاصية')
    return redirect('case_detail', pk=pk)


@login_required
def case_type_view(request, case_type):
    qs = Case.objects.select_related('beneficiary', 'assigned_to').filter(case_type=case_type)
    labels = dict(Case.CASE_TYPES)
    type_label = labels.get(case_type, case_type)
    total = qs.count()
    active = qs.filter(status__in=['new', 'under_study', 'approved']).count()
    completed = qs.filter(status__in=['disbursed', 'closed']).count()
    urgent = qs.filter(priority='urgent').count()
    total_amount = qs.aggregate(total=Sum('requested_amount'))['total'] or 0
    return render(request, 'cases/type_detail.html', {
        'cases': qs.order_by('-created_at'),
        'case_type': case_type,
        'type_label': type_label,
        'total': total, 'active': active,
        'completed': completed, 'urgent': urgent,
        'total_amount': total_amount,
    })


@login_required
def case_type_export_excel(request, case_type):
    labels = dict(Case.CASE_TYPES)
    type_label = labels.get(case_type, case_type)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = type_label
    headers = ['رقم الحالة', 'المستفيد', 'النوع', 'الأولوية', 'المبلغ المطلوب',
               'المبلغ المعتمد', 'الحالة', 'تاريخ الإنشاء']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row, c in enumerate(Case.objects.select_related('beneficiary').filter(case_type=case_type), 2):
        vals = [c.code, c.beneficiary.full_name if c.beneficiary else '',
                c.get_case_type_display(), c.get_priority_display(),
                float(c.requested_amount or 0), float(c.approved_amount or 0),
                c.get_status_display(), c.created_at.strftime('%Y-%m-%d')]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={case_type}_cases.xlsx'
    wb.save(response)
    return response


@login_required
def case_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الحالات'
    headers = ['رقم الحالة', 'المستفيد', 'النوع', 'الأولوية', 'المبلغ المطلوب',
               'المبلغ المعتمد', 'الحالة', 'تاريخ الإنشاء']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row, c in enumerate(Case.objects.select_related('beneficiary').all(), 2):
        vals = [c.code, c.beneficiary.full_name if c.beneficiary else '',
                c.get_case_type_display(), c.get_priority_display(),
                c.requested_amount or 0, c.approved_amount or 0,
                c.get_status_display(), c.created_at.strftime('%Y-%m-%d')]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cases.xlsx'
    wb.save(response)
    return response
